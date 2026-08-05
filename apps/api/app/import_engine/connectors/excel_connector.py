"""Excel (.xlsx) connector via openpyxl."""

from __future__ import annotations

import io
from collections import defaultdict
from typing import Any

from app.import_engine.connectors.base import ConnectorContext
from app.import_engine.connectors.csv_connector import ENTITY_ALIASES
from app.import_engine.enums import ImportSource
from app.import_engine.models import NormalizedBatch
from app.import_engine.normalize import build_batch_from_sections

_KNOWN_SHEETS = set(ENTITY_ALIASES.values())


class ExcelConnector:
    source = ImportSource.EXCEL

    async def extract(self, ctx: ConnectorContext) -> NormalizedBatch:
        if not ctx.payload:
            raise ValueError("Excel import requires an uploaded file")
        name = (ctx.filename or "").lower()
        if name.endswith((".xls", ".xlt")) and not name.endswith((".xlsx", ".xlsm", ".xltx")):
            raise ValueError("Legacy .xls is not supported — save as .xlsx and retry")
        # ZIP/OOXML signature — reject CSV/text accidentally labeled as Excel
        if not ctx.payload.startswith(b"PK"):
            raise ValueError(
                "File is not a valid .xlsx workbook. If this is a CSV, choose CSV import."
            )
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("openpyxl is required for Excel import") from exc

        try:
            wb = load_workbook(io.BytesIO(ctx.payload), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 — surface user-facing parse errors
            raise ValueError(f"Could not read Excel file: {exc}") from exc

        sections: dict[str, list[dict[str, Any]]] = defaultdict(list)
        try:
            for sheet in wb.worksheets:
                sheet_key = ENTITY_ALIASES.get(
                    sheet.title.lower().strip(), sheet.title.lower().strip()
                )
                rows = sheet.iter_rows(values_only=True)
                try:
                    header = next(rows)
                except StopIteration:
                    continue
                headers = [
                    str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header)
                ]
                if not any(headers):
                    continue
                for idx, values in enumerate(rows, start=2):
                    if values is None or all(v is None or str(v).strip() == "" for v in values):
                        continue
                    row = {
                        headers[i]: ("" if v is None else str(v).strip())
                        for i, v in enumerate(values)
                        if i < len(headers) and headers[i]
                    }
                    row["row_ref"] = f"excel:{sheet.title}:{idx}"
                    entity = sheet_key if sheet_key in _KNOWN_SHEETS else "customers"
                    for key in ("entity", "entity_type", "type", "record_type"):
                        if key in row and row[key]:
                            entity = ENTITY_ALIASES.get(str(row.pop(key)).lower(), entity)
                            break
                    if sheet_key in _KNOWN_SHEETS:
                        entity = sheet_key
                    sections[entity].append(row)
        finally:
            wb.close()

        if not sections:
            raise ValueError("Excel workbook contained no data rows")
        return build_batch_from_sections(dict(sections), source=self.source)
