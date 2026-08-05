"""Phase 9 Import Engine tests."""

from __future__ import annotations

from uuid import uuid4

import pytest

from app.agents.factory import build_agent_runtime
from app.import_engine.enums import ImportJobStatus, ImportSource, MergeAction
from app.import_engine.factory import build_import_runtime, reset_import_runtime
from app.import_engine.store import InMemoryImportStore
from app.import_engine.vin import validate_vin


@pytest.fixture(autouse=True)
def _reset():
    reset_import_runtime()
    yield
    reset_import_runtime()


@pytest.fixture
def shop_id():
    return uuid4()


@pytest.fixture
def runtime():
    # Explicit in-memory agents so unit tests do not require Postgres CRM.
    return build_import_runtime(store=InMemoryImportStore(), agents=build_agent_runtime())


@pytest.mark.asyncio
async def test_tekmetric_sample_import_auto_apply(runtime, shop_id):
    job = await runtime.service.create_job(
        shop_id=shop_id,
        source=ImportSource.TEKMETRIC,
        options={"use_sample": True},
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert job.status == ImportJobStatus.COMPLETED
    assert job.report is not None
    assert job.report.entity_counts["customer"].imported >= 1
    assert job.report.entity_counts["vehicle"].imported >= 1
    assert job.report.entity_counts["repair_history"].imported >= 1
    assert job.report.entity_counts["invoice"].imported >= 1


@pytest.mark.asyncio
async def test_csv_import_and_normalize(runtime, shop_id):
    from datetime import datetime, timezone

    # Simpler dedicated CSV with entity column per row type
    csv_body = b"""entity,name,phone,email,vin,year,make,model,mileage,service_type,description,cost,date
customer,Pat Kim,555-2222,pat@example.com,,,,,,,,,
vehicle,,,,1HGCM82633A004352,2019,Toyota,Camry,40000,,,,,
repair,,,,1HGCM82633A004352,,,,,oil_change,Oil & filter,89.99,2025-11-12
"""
    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.CSV)
    await runtime.service.attach_upload(
        shop_id=shop_id, job_id=job.id, payload=csv_body, filename="shop.csv"
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert job.status == ImportJobStatus.COMPLETED
    assert job.batch is not None
    assert job.batch.counts()["customer"] == 1
    assert job.batch.counts()["vehicle"] == 1
    assert job.batch.counts()["repair_history"] == 1
    assert job.batch.repairs[0].performed_at == datetime(2025, 11, 12, tzinfo=timezone.utc)
    # Apply must land in agent CRM directories (customers UI source when SQL-backed).
    customers = await runtime.agents.customer.directory.find_by_phone(shop_id, "5552222")
    assert len(customers) == 1
    assert customers[0].name == "Pat Kim"
    vehicle = await runtime.agents.vehicle.directory.find_by_vin(shop_id, "1HGCM82633A004352")
    assert vehicle is not None
    assert vehicle.make == "Toyota"
    repairs = await runtime.agents.vehicle.directory.list_repairs(shop_id, vehicle.id)
    assert len(repairs) == 1
    assert repairs[0].performed_at == datetime(2025, 11, 12, tzinfo=timezone.utc)


@pytest.mark.asyncio
async def test_repair_without_vehicle_is_not_counted_imported(runtime, shop_id):
    csv_body = b"""entity,name,phone,email,vin,year,make,model,mileage,service_type,description,cost,date
customer,Pat Kim,555-2222,pat@example.com,,,,,,,,,
repair,,,,1HGCM82633A004352,,,,,oil_change,Oil & filter,89.99,2025-11-12
"""
    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.CSV)
    await runtime.service.attach_upload(
        shop_id=shop_id, job_id=job.id, payload=csv_body, filename="orphan_repair.csv"
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert job.status == ImportJobStatus.COMPLETED
    assert job.report is not None
    assert job.report.entity_counts["repair_history"].imported == 0
    assert job.report.entity_counts["repair_history"].failed >= 1

@pytest.mark.asyncio
async def test_excel_import_and_normalize(runtime, shop_id):
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "customers"
    ws.append(["name", "phone", "email"])
    ws.append(["Pat Kim", "555-2222", "pat@example.com"])
    ws2 = wb.create_sheet("vehicles")
    ws2.append(["vin", "year", "make", "model", "mileage"])
    ws2.append(["1HGCM82633A004352", 2019, "Toyota", "Camry", 40000])
    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.EXCEL)
    await runtime.service.attach_upload(
        shop_id=shop_id, job_id=job.id, payload=payload, filename="shop.xlsx"
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert job.status == ImportJobStatus.COMPLETED
    assert job.batch is not None
    assert job.batch.counts()["customer"] == 1
    assert job.batch.counts()["vehicle"] == 1


@pytest.mark.asyncio
async def test_xlsx_uploaded_as_csv_source_is_rewired(runtime, shop_id):
    """UI often defaults to CSV; uploading .xlsx must switch to the Excel connector."""
    import io

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "customers"
    ws.append(["name", "phone", "email"])
    ws.append(["Sam Chen", "555-0199", "sam@example.com"])
    buf = io.BytesIO()
    wb.save(buf)
    payload = buf.getvalue()

    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.CSV)
    job = await runtime.service.attach_upload(
        shop_id=shop_id, job_id=job.id, payload=payload, filename="export.xlsx"
    )
    assert job.source == ImportSource.EXCEL
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert job.status == ImportJobStatus.COMPLETED
    assert job.batch is not None
    assert job.batch.counts()["customer"] == 1


@pytest.mark.asyncio
async def test_duplicate_customers_await_resolution(runtime, shop_id):
    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.MANUAL)
    await runtime.service.set_manual_sections(
        shop_id=shop_id,
        job_id=job.id,
        sections={
            "customers": [
                {"name": "Alex Rivera", "phone": "555-0100", "email": "a@example.com"},
                {"name": "Alex R", "phone": "555-0100", "email": "alex@example.com"},
            ]
        },
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=False)
    assert job.status == ImportJobStatus.AWAITING_RESOLUTION
    assert any(d.match_type.value == "phone" for d in job.duplicates)

    resolutions = [
        {"duplicate_id": str(job.duplicates[0].id), "action": MergeAction.MERGE.value}
    ]
    # resolve all pending
    resolutions = [
        {"duplicate_id": str(d.id), "action": MergeAction.MERGE.value}
        for d in job.duplicates
        if not d.resolved
    ]
    job = await runtime.service.resolve_duplicates(
        shop_id=shop_id, job_id=job.id, resolutions=resolutions, apply_after=True
    )
    assert job.status == ImportJobStatus.COMPLETED
    assert job.report is not None
    assert job.report.entity_counts["customer"].imported == 1


@pytest.mark.asyncio
async def test_invalid_vin_detected(runtime, shop_id):
    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.MANUAL)
    await runtime.service.set_manual_sections(
        shop_id=shop_id,
        job_id=job.id,
        sections={
            "vehicles": [
                {
                    "vin": "1HGCM82633A004353",  # bad check digit
                    "year": 2018,
                    "make": "Honda",
                    "model": "Accord",
                    "mileage": 10,
                }
            ]
        },
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert any(i.code == "invalid_vin" for i in job.validation_issues)


@pytest.mark.asyncio
async def test_inconsistent_mileage_detected(runtime, shop_id):
    job = await runtime.service.create_job(shop_id=shop_id, source=ImportSource.MANUAL)
    await runtime.service.set_manual_sections(
        shop_id=shop_id,
        job_id=job.id,
        sections={
            "vehicles": [
                {
                    "vin": "1HGCM82633A004352",
                    "year": 2018,
                    "make": "Honda",
                    "model": "Accord",
                    "mileage": 80000,
                }
            ],
            "repairs": [
                {
                    "vin": "1HGCM82633A004352",
                    "service_type": "oil_change",
                    "description": "oil",
                    "mileage": 72000,
                    "date": "2025-06-01",
                    "cost": 89,
                },
                {
                    "vin": "1HGCM82633A004352",
                    "service_type": "brakes",
                    "description": "pads",
                    "mileage": 50000,
                    "date": "2025-12-01",
                    "cost": 200,
                },
            ],
        },
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert any(i.code == "inconsistent_mileage" for i in job.validation_issues)


@pytest.mark.asyncio
async def test_ocr_text_extraction(runtime, shop_id):
    text = """
    Customer: Casey Morgan
    Phone: 555-4444
    VIN: 1HGCM82633A004352
    2018 Honda Accord
    Mileage: 55,200
    Repair: Timing belt replacement
    Invoice #: INV-900
    Total: $1,240.00
    Recommendation: Replace water pump with belt
    """
    job = await runtime.service.create_job(
        shop_id=shop_id, source=ImportSource.OCR, options={"ocr_text": text}
    )
    job = await runtime.service.run(shop_id=shop_id, job_id=job.id, auto_apply=True)
    assert job.status == ImportJobStatus.COMPLETED
    assert job.batch is not None
    assert job.batch.customers[0].name.startswith("Casey")
    assert job.batch.vehicles[0].vin == "1HGCM82633A004352"
    assert job.batch.invoices
    assert job.batch.recommendations


def test_vin_check_digit():
    ok, err = validate_vin("1HGCM82633A004352")
    assert ok and err is None
    ok, err = validate_vin("1HGCM82633A004353")
    assert not ok
